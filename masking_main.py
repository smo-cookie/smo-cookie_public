import os
import json
import re
import openai
import zipfile
from pymongo import MongoClient
from docx import Document
from openpyxl import load_workbook
from tempfile import TemporaryDirectory
from lxml import etree

# openai api
client = openai.OpenAI(api_key="")

# 연결 -> 몽고디비
MONGO_URI = "mongodb+srv://smocookie:smocookie@cluster0.btwrt.mongodb.net/?retryWrites=true&w=majority"
client_db = MongoClient(MONGO_URI)
db = client_db["personal_info_db"]
detected_info_collection = db["detected_info"]
file_metadata_collection = db["file_metadata"]
additional_info_collection = db["additional_info"]

MASKED_DIR = "masked_files"
if not os.path.exists(MASKED_DIR):  
    os.makedirs(MASKED_DIR)

# 정규표현식
patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    "연락처": r"\b010-\d{4}-\d{4}\b",
    "생년월일": r"\b\d{4}[-/]\d{2}[-/]\d{2}\b",
    "계좌번호": r"\b\d{2,4}-\d{2,4}-\d{2,4}\b",
    "여권번호": r"\b[A-Z]{1}\d{8}\b",
    "이메일": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    "카드번호": r"\b\d{4}-\d{4}-\d{4}-\d{4}\b"
}

# 정규표현식 기반 탐지
def detect_pii_with_regex(content, selected_types):
    results = {}
    for key, pattern in patterns.items():
        if key in selected_types:
            matches = re.findall(pattern, content)
            if matches:
                results[key] = list(set(matches))  # 중복 제거
    return results

# open ai 기반탐지
def detect_sensitive_info_with_chatgpt(content, selected_types, additional_info):
    prompt = f"""
    다음 텍스트에서 다음 항목을 문맥을 분석하여 탐지하세요:
    - 선택된 개인정보: {selected_types}
    - 추가 요청 정보: {additional_info}
    
    **반환 형식(JSON):**
    **아래는 예시입니다. 이런 식의 정보들을 찾아서 반환하면 됩니다."
    {{
        "개인정보": {{
            "이름": ["홍길동", "김철수"],
            "주소": ["서울시 강남구 역삼동"]
        }},
        "추가 탐지 정보": {{
            "추가 요청 정보": ["Project Alpha", "XYZ Corporation"]
        }}
    }}

    **분석할 텍스트:**
    {content}
    """

    response = client.chat.completions.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}],
    )

    try:
        return json.loads(response.choices[0].message.content)
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from ChatGPT"}

# 디비에 탐지된 정보 저장
def save_to_mongodb(file_name, detected_info, additional_results):
    file_metadata_collection.insert_one({"file_name": file_name})
    detected_info_collection.insert_one({"file_name": file_name, "detected_info": detected_info})
    additional_info_collection.insert_one({"file_name": file_name, "additional_info": additional_results})

# 마스킹 데이터 가져오기
def get_masking_data_from_mongodb(file_name):
    masking_data = set()
    detected_info = detected_info_collection.find_one({"file_name": file_name})
    if detected_info and "detected_info" in detected_info:
        for values in detected_info["detected_info"].values():
            masking_data.update(values)

    additional_info = additional_info_collection.find_one({"file_name": file_name})
    if additional_info and "additional_info" in additional_info:
        masking_data.update(additional_info["additional_info"])

    return masking_data

# 마스킹 적용
def apply_masking(content, masking_data):
    for item in masking_data:
        content = content.replace(item, "****")
    return content

# 워드 파일 텍스트 추출
def extract_text_from_word(file_path):
    document = Document(file_path)
    return "\n".join([paragraph.text for paragraph in document.paragraphs])

# 엑셀 텍스트 추출
def extract_text_from_excel(file_path):
    workbook = load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text

# xml 기반 마스킹 적용
def process_xml_file(xml_path, masking_data):
    parser = etree.XMLParser(remove_blank_text=True)
    with open(xml_path, 'rb') as file:
        xml_tree = etree.parse(file, parser)

    for element in xml_tree.xpath("//w:t", namespaces={"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        if element.text:
            element.text = apply_masking(element.text, masking_data)

    with open(xml_path, 'wb') as file:
        file.write(etree.tostring(xml_tree, pretty_print=True))

# Word 파일 마스킹
def mask_sensitive_data_with_images(file_path):
    masking_data = get_masking_data_from_mongodb(file_path)
    if not masking_data:
        print("No data to mask")
        return None

    with TemporaryDirectory() as temp_dir:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        document_xml_path = os.path.join(temp_dir, "word", "document.xml")
        if os.path.exists(document_xml_path):
            process_xml_file(document_xml_path, masking_data)

        file_name = os.path.basename(file_path).replace(".docx", "(masked).docx")
        masked_file_path = os.path.join(MASKED_DIR, file_name)
        with zipfile.ZipFile(masked_file_path, 'w') as zip_out:
            for foldername, subfolders, filenames in os.walk(temp_dir):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_out.write(file_path, arcname)

    return masked_file_path if os.path.exists(masked_file_path) else None

# main
def main(file_path, file_type, selected_types_json, additional_info_json):
    selected_types = json.loads(selected_types_json) if selected_types_json else []
    additional_info = json.loads(additional_info_json) if additional_info_json else []
    

    if file_type == "word":
        content = extract_text_from_word(file_path)
    elif file_type == "excel":
        content = extract_text_from_excel(file_path)
    else:
        print("지원하지 않는 파일 형식입니다.")
        return None

    regex_results = detect_pii_with_regex(content, selected_types)
    chatgpt_response = detect_sensitive_info_with_chatgpt(content, selected_types, additional_info)

    if "error" in chatgpt_response:
        print("ChatGPT 탐지 중 오류 발생:", chatgpt_response["error"])
        return None

    final_results = {**regex_results, **chatgpt_response.get("개인정보", {})}

    save_to_mongodb(file_path, final_results, chatgpt_response.get("추가 탐지 정보", {}))

    masked_file = mask_sensitive_data_with_images(file_path)

    if not masked_file or not os.path.exists(masked_file):
        print("마스킹된 파일이 생성되지 않았습니다.")
        return None  

    print(f"마스킹된 파일이 저장되었습니다: {masked_file}")
    return masked_file  # 파일 경로 반환
